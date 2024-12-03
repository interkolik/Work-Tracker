import datetime
import pandas as pd
import os
import shutil
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import pytz  # Importing timezone library

# Set the PST timezone
pst = pytz.timezone('US/Pacific')

def create_template(start_date):
    wb = Workbook()
    ws = wb.active
    ws.title = start_date.strftime('%d %B %y')  # Rename initial sheet

    ws.append(['Date'] + [(start_date + datetime.timedelta(days=i)).strftime('%d.%m.%y') for i in range(7)])
    ws.append(['Day'] + ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday'])
    for hour in range(24):
        for minute in range(0, 60, 15):
            time_label = f'{hour:02}-{minute:02}'
            row = [time_label] + ['' for _ in range(7)]
            ws.append(row)

    for col_num in range(2, 9):
        col_letter = get_column_letter(col_num)
        ws.column_dimensions[col_letter].width = 35

    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border

    wb.save('work_log.xlsx')

def backup_file(file_path):
    backup_dir = os.path.join(os.path.dirname(file_path), 'backup')
    if not os.path.exists(backup_dir):
        os.makedirs(backup_dir)
    backup_file_path = os.path.join(backup_dir, f'work_log_backup_{datetime.datetime.now(pst).strftime("%Y%m%d%H%M%S")}.xlsx')
    shutil.copy(file_path, backup_file_path)
    print(f"Backup created: {backup_file_path}")

def create_or_merge_ics_file(summary, start_time_utc, end_time_utc, tasks):
    calendar_dir = os.path.join(os.getcwd(), 'calendar')
    if not os.path.exists(calendar_dir):
        os.makedirs(calendar_dir)
    
    filename = os.path.join(calendar_dir, "merged_tasks.ics")
    
    new_event = (
        "BEGIN:VEVENT\n"
        f"UID:{datetime.datetime.utcnow().strftime('%Y%m%dT%H%M%SZ')}@yourdomain.com\n"
        f"DTSTAMP:{datetime.datetime.utcnow().strftime('%Y%m%dT%H%M%SZ')}\n"
        f"DTSTART:{start_time_utc.strftime('%Y%m%dT%H%M%SZ')}\n"
        f"DTEND:{end_time_utc.strftime('%Y%m%dT%H%M%SZ')}\n"
        f"SUMMARY:{summary}\n"
        "END:VEVENT\n"
    )
    
    # Read existing content if file exists
    if os.path.exists(filename):
        with open(filename, 'r') as file:
            content = file.read()
        content = content.replace("END:VCALENDAR\n", new_event + "END:VCALENDAR\n")
    else:
        # If file doesn't exist, create new content
        content = (
            "BEGIN:VCALENDAR\n"
            "VERSION:2.0\n"
            "PRODID:-//Your Organization//Your Product//EN\n"
            + new_event +
            "END:VCALENDAR\n"
        )
    
    # Write back the merged content
    with open(filename, 'w') as file:
        file.write(content)
    
    print(f"ICS entry added: {filename}")
    
def record_time(job_title, tasks, start_time, end_time):
    # Convert the start and end times to PST
    start_time = start_time.astimezone(pst)
    end_time = end_time.astimezone(pst)

    time_worked = end_time - start_time
    week_start = start_time - datetime.timedelta(days=start_time.weekday())  # Start of the week
    sheet_title = week_start.strftime('%d %B %y')
    start_time_str = start_time.strftime('%H:%M')
    end_time_str = end_time.strftime('%H:%M')

    print(f"Recording time: start_time={start_time_str}, end_time={end_time_str}")  # Debugging print

    if not os.path.isfile('work_log.xlsx'):
        create_template(start_date=week_start)
    else:
        backup_file('work_log.xlsx')

    wb = load_workbook('work_log.xlsx')
    if sheet_title not in wb.sheetnames:
        ws = wb.create_sheet(title=sheet_title)
        ws.append(['Date'] + [(week_start + datetime.timedelta(days=i)).strftime('%d.%m.%y') for i in range(7)])
        ws.append(['Day'] + ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday'])
        for hour in range(24):
            for minute in range(0, 60, 15):
                time_label = f'{hour:02}-{minute:02}'
                row = [time_label] + ['' for _ in range(7)]
                ws.append(row)
        
        for col_num in range(2, 9):
            col_letter = get_column_letter(col_num)
            ws.column_dimensions[col_letter].width = 35

        thin_border = Border(left=Side(style='thin'), 
                             right=Side(style='thin'), 
                             top=Side(style='thin'), 
                             bottom=Side(style='thin'))

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = thin_border
    else:
        ws = wb[sheet_title]

    day_index = start_time.weekday() + 2  # 0 = Monday; 2nd column in sheet
    start_row = int(start_time.strftime('%H')) * 4 + int(start_time.strftime('%M')) // 15 + 3
    end_row = int(end_time.strftime('%H')) * 4 + int(end_time.strftime('%M')) // 15 + 3

    if ws.cell(row=start_row, column=day_index).value:
        ws.cell(row=start_row, column=day_index).fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow for override
    ws.cell(row=start_row, column=day_index).value = f"{job_title}, {tasks}"

    if start_row != end_row:
        for r in range(start_row + 1, end_row):
            ws.cell(row=r, column=day_index).value = "Working"
            ws.cell(row=r, column=day_index).fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        ws.cell(row=end_row, column=day_index).value = "Done"
        ws.cell(row=end_row, column=day_index).fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

    date_sheets = [sheet for sheet in wb.sheetnames if not sheet.startswith('Sheet')]
    wb._sheets = [wb[sheet] for sheet in sorted(date_sheets, key=lambda sheet: datetime.datetime.strptime(sheet, '%d %B %y'), reverse=True)]

    wb.save('work_log.xlsx')

    # Convert local time to UTC for .ics file
    start_time_utc = start_time.astimezone(pytz.utc)
    end_time_utc = end_time.astimezone(pytz.utc)

    print(f"Creating ICS: start_time_utc={start_time_utc.strftime('%Y%m%dT%H%M%SZ')}, end_time_utc={end_time_utc.strftime('%Y%m%dT%H%M%SZ')}")
    # Debugging print
    create_or_merge_ics_file(job_title, start_time_utc, end_time_utc, tasks)

def main():
    while True:
        command = input("Enter 'start' to begin tracking time and 'stop' to end tracking: ").strip().lower()
        if command == 'start':
            start_time = datetime.datetime.now(pst)
            print(f"Started tracking at {start_time}")
        elif command == 'stop':
            end_time = datetime.datetime.now(pst)
            print(f"Stopped tracking at {end_time}")
            job_title = input("Enter the job title: ").strip()
            tasks = input("Enter number of tasks completed: ").strip()
            record_time(job_title, tasks, start_time, end_time)
            print("Work time and tasks recorded.")
            break
        elif command == 'manual':
            start_time_str = input("Enter the start time (YYYY-MM-DD HH:MM PST): ").strip()
            end_time_str = input("Enter the end time (YYYY-MM-DD HH:MM PST): ").strip()
            start_time = pst.localize(datetime.datetime.strptime(start_time_str, '%Y-%m-%d %H:%M'))
            end_time = pst.localize(datetime.datetime.strptime(end_time_str, '%Y-%m-%d %H:%M'))
            job_title = input("Enter the job title: ").strip()
            tasks = input("Enter number of tasks completed: ").strip()
            record_time(job_title, tasks, start_time, end_time)
            print("Work time and tasks recorded manually.")
            break
        else:
            print("Invalid command. Please enter 'start' or 'stop'.")

if __name__ == "__main__":
    main()


